from openpyxl import load_workbook, Workbook
import re

src = "６年・５年度売上比較_新_renamed.xlsx"

wb = load_workbook(src, data_only=True)

# 出力ブック
wb_keep_abe_fh = Workbook()
wb_keep_cdgi_prev = Workbook()

# 既定の最初の空シートを削除
wb_keep_abe_fh.remove(wb_keep_abe_fh.active)
wb_keep_cdgi_prev.remove(wb_keep_cdgi_prev.active)

# ヘルパー：全角→半角
def to_halfwidth_num(s: str) -> str:
    return s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))

# ヘルパー：シート名（YYYY-M or YYYY-MM）から (year, month) を取得
def parse_year_month(title: str):
    t = to_halfwidth_num(title.strip())
    m = re.search(r"^\s*(\d{4})\s*[-_/．\.]\s*(\d{1,2})\s*$", t)
    if m:
        return int(m.group(1)), int(m.group(2))
    # 年のみ
    m2 = re.search(r"^\s*(\d{4})\s*$", t)
    if m2:
        return int(m2.group(1)), None
    return None, None

# 列のインデックス取得（A=1, B=2 ...）
def col_idx(letter: str) -> int:
    total = 0
    for ch in letter.upper():
        if 'A' <= ch <= 'Z':
            total = total * 26 + (ord(ch) - ord('A') + 1)
    return total

cols_left_1 = ['A','B','E','F','H']
cols_left_2 = ['C','D','G','I']

for ws in wb.worksheets:
    # --- 1) A,B,E,F,H を残す ---
    new_ws1 = wb_keep_abe_fh.create_sheet(title=ws.title)
    # 書き込み
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        new_row = []
        for col_letter in cols_left_1:
            ci = col_idx(col_letter)
            if ci <= ws.max_column:
                new_row.append(r[ci-1].value)
            else:
                new_row.append(None)
        new_ws1.append(new_row)
    
    # --- 2) C,D,G,I を残す（シート名は1年前の西暦へ） ---
    year, month = parse_year_month(ws.title)
    if year is not None:
        prev_year = year - 1
        if month is not None:
            new_title = f"{prev_year}-{month:02d}"
        else:
            new_title = f"{prev_year}"
    else:
        # パースできない場合は元名 + "_prev" にフォールバック
        new_title = ws.title + "_prev"
    
    # シート名重複回避
    base_title = new_title
    suffix = 1
    while new_title in wb_keep_cdgi_prev.sheetnames:
        new_title = f"{base_title}_{suffix}"
        suffix += 1
    
    new_ws2 = wb_keep_cdgi_prev.create_sheet(title=new_title)
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        new_row = []
        for col_letter in cols_left_2:
            ci = col_idx(col_letter)
            if ci <= ws.max_column:
                new_row.append(r[ci-1].value)
            else:
                new_row.append(None)
        new_ws2.append(new_row)

out1 = "６年・５年度売上比較_新_ABEFH.xlsx"
out2 = "６年・５年度売上比較_新_CDGI_prevyear.xlsx"
wb_keep_abe_fh.save(out1)
wb_keep_cdgi_prev.save(out2)

(out1, out2)
