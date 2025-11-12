from openpyxl import load_workbook
import re

# === 設定 ===
src_path = "６年・５年度売上比較_新.xlsx"
dst_path = "６年・５年度売上比較_新_renamed.xlsx"

# 和暦変換用
REIWA_START = 2018  # 令和1年=2019年 → 西暦 = REIWA_START + n

# ブック読み込み
wb = load_workbook(src_path)

for ws in wb.worksheets:
    old_name = ws.title.strip()

    # 「〇年〇月」パターンを探す（例：「5年4月」「６年１２月」）
    m = re.search(r"([０-９0-9]+)\s*年\s*([０-９0-9]+)\s*月", old_name)
    if not m:
        continue

    # 和数字→半角数字へ
    year_str = m.group(1).translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    month_str = m.group(2).translate(str.maketrans("０１２３４５６７８９", "0123456789"))

    reiwa_year = int(year_str)
    month = int(month_str)

    # 西暦へ変換
    western_year = REIWA_START + reiwa_year

    # 新しいシート名（例：2023-4）
    new_name = f"{western_year}-{month}"

    # シート名が重複していたら "_dup" をつける
    if new_name in wb.sheetnames:
        new_name += "_dup"

    print(f"{old_name} → {new_name}")
    ws.title = new_name

# 保存
wb.save(dst_path)
print(f"完了：{dst_path}")
